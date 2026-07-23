"""Fills the DPIA Excel risk register from saved structured data."""

import os
import openpyxl

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "MCP Risks.xlsx")

# Rows in the template where risk data starts (1-indexed, row 3 = first data row)
FIRST_DATA_ROW = 3
MAX_RISKS = 11  # template has rows 3-13

# Column indices (1-indexed)
COL_RISK        = 2   # B: Risk name
COL_DESC        = 3   # C: Description
COL_PROB_BEFORE = 4   # D: Probability before mitigation
COL_SEV_BEFORE  = 5   # E: Severity before mitigation
COL_ASSESS_BEFORE = 6 # F: Risk assessment before
COL_MEASURE     = 7   # G: Mitigation measure
COL_PROB_AFTER  = 8   # H: Probability after mitigation
COL_SEV_AFTER   = 9   # I: Severity after mitigation
COL_ASSESS_AFTER = 10 # J: Risk assessment after


def fill_risks(saved: dict) -> str:
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    steg4 = saved.get("steg4", {})
    risks = steg4.get("risks", []) if isinstance(steg4, dict) else []

    for i, risk in enumerate(risks[:MAX_RISKS]):
        row = FIRST_DATA_ROW + i
        ws.cell(row, COL_RISK).value        = risk.get("risk", "")
        ws.cell(row, COL_DESC).value        = risk.get("beskrivning", "")
        ws.cell(row, COL_PROB_BEFORE).value = risk.get("sannolikhet", "")
        ws.cell(row, COL_SEV_BEFORE).value  = risk.get("konsekvensniva", "")
        ws.cell(row, COL_ASSESS_BEFORE).value = risk.get("riskbedomning", "")
        ws.cell(row, COL_MEASURE).value     = risk.get("atgard", "")
        ws.cell(row, COL_PROB_AFTER).value  = risk.get("sannolikhet_efter", "")
        ws.cell(row, COL_SEV_AFTER).value   = risk.get("konsekvensniva_efter", "")
        ws.cell(row, COL_ASSESS_AFTER).value = risk.get("riskbedomning_efter", "")

    title = saved.get("title", "dpia")
    safe = title.lower().replace(" ", "_").replace("/", "_")
    out_dir = os.path.join(os.path.dirname(__file__), "output")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"{safe}_risks.xlsx")
    wb.save(out_path)
    return out_path
